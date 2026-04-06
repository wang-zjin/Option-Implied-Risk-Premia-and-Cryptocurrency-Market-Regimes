"""
CL23 分区域 / 全样本 BP、BVRP：对外生变量做「变化对变化」回归（仅 CL23）。

区域列来自 CL23 工作簿（**BP_BVRP_CL23.xlsx**，见 S14 路径解析；子表 BP、VRP，日期列 date）：
  • UP：**rp_up**  |  Down：**rp_down**  |  Total：**rp_global_sum**

每个区域：因变量 ΔBP_{t+1}、ΔBVRP_{t+1}（forward 差分），自变量 X_t−X_{t-1}；BVRP 符号与
S14 中 CL23 一致（negate_bvrp_for_display=False）。

结果目录 S14_1_regress_explain/CL23/：
  • UP：UP_region_*.csv
  • Down：Down_region_*.csv
  • Total：Total_region_*.csv
"""

from __future__ import annotations

import math
import os
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd

from S14_correlation import (
    BASE_DIR,
    BOUNDS_RUNS,
    CL23_BP_BVRP_XLSX,
    BoundsRunConfig,
    ensure_derived_observables,
    list_observable_files,
    load_observable,
    merge_for_plot,
    observable_plot_label,
    observable_value_columns,
    _parse_yyyymmdd_or_datetime,
)

# 仅 CL23 含工作簿内 UP 区域（rp_up）；不跑 CL20 / SS25。
REGRESSION_RUNS: tuple[BoundsRunConfig, ...] = tuple(
    r for r in BOUNDS_RUNS if r.key == "CL23"
)

OUTPUT_ROOT = BASE_DIR / "S14_1_regress_explain"

RESULT_COLUMNS: tuple[str, ...] = (
    "bounds_key",
    "region",
    "dependent",
    "observable_file",
    "observable_column",
    "observable_label",
    "n",
    "intercept",
    "slope",
    "se_slope",
    "t statistics",
    "p value",
    "r2",
    "note",
)

SUMMARY_COLUMNS: tuple[str, ...] = (
    "dependent",
    "observable_label",
    "observable_file",
    "observable_column",
    "n",
    "slope",
    "t statistics",
    "p value",
    "stars",
    "sig_level",
    "r2",
)


def _as_float_p(p: object) -> float:
    try:
        x = float(p)
    except (TypeError, ValueError):
        return float("nan")
    return x


def _sig_stars(p: object) -> str:
    pp = _as_float_p(p)
    if math.isnan(pp):
        return ""
    if pp < 0.01:
        return "***"
    if pp < 0.05:
        return "**"
    if pp < 0.10:
        return "*"
    return ""


def _sig_level(p: object) -> str:
    pp = _as_float_p(p)
    if math.isnan(pp):
        return ""
    if pp < 0.01:
        return "1%"
    if pp < 0.05:
        return "5%"
    if pp < 0.10:
        return "10%"
    return "n.s."


def _load_cl23_region_from_bp_bvrp_xlsx(
    path: Path,
    rp_col: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """从 BP / VRP 子表读取同一风险溢价列（如 rp_up、rp_down、rp_global_sum）。"""
    key = rp_col.strip().lower()
    bp_raw = pd.read_excel(path, sheet_name="BP", engine="openpyxl")
    vrp_raw = pd.read_excel(path, sheet_name="VRP", engine="openpyxl")
    bp_raw.columns = [str(c).strip() for c in bp_raw.columns]
    vrp_raw.columns = [str(c).strip() for c in vrp_raw.columns]
    cl_b = {c.lower(): c for c in bp_raw.columns}
    cl_v = {c.lower(): c for c in vrp_raw.columns}
    d_b = cl_b.get("date")
    d_v = cl_v.get("date")
    if d_b is None or key not in cl_b:
        raise ValueError(
            f"{path.name} 表 BP 需要列 date 与 {rp_col}；当前: {list(bp_raw.columns)}"
        )
    if d_v is None or key not in cl_v:
        raise ValueError(
            f"{path.name} 表 VRP 需要列 date 与 {rp_col}；当前: {list(vrp_raw.columns)}"
        )
    dt_b = _parse_yyyymmdd_or_datetime(bp_raw[d_b])
    dt_v = _parse_yyyymmdd_or_datetime(vrp_raw[d_v])
    bp = pd.DataFrame(
        {
            "Date": dt_b,
            "BP": pd.to_numeric(bp_raw[cl_b[key]], errors="coerce"),
        }
    ).dropna(subset=["Date"])
    bvrp = pd.DataFrame(
        {
            "Date": dt_v,
            "BVRP": pd.to_numeric(vrp_raw[cl_v[key]], errors="coerce"),
        }
    ).dropna(subset=["Date"])
    return bp, bvrp


def load_cl23_region_bounds(
    run: BoundsRunConfig,
    rp_col: str,
) -> tuple[pd.DataFrame, pd.DataFrame] | None:
    if run.key != "CL23":
        return None
    cl23_path = run.combined_bp_bvrp_path or CL23_BP_BVRP_XLSX
    if not cl23_path.is_file():
        return None
    try:
        return _load_cl23_region_from_bp_bvrp_xlsx(cl23_path, rp_col)
    except Exception:
        return None


@dataclass(frozen=True)
class RegionRunSpec:
    """file_prefix 用于输出文件名，如 UP_region、Down_region、Total_region。"""

    file_prefix: str
    rp_excel_col: str
    region_code: str
    dep_bp: str
    dep_bvrp: str
    write_summary_xlsx: bool


CL23_REGION_SPECS: tuple[RegionRunSpec, ...] = (
    RegionRunSpec("UP_region", "rp_up", "UP", "UP_BP", "UP_BVRP", True),
    RegionRunSpec("Down_region", "rp_down", "DOWN", "DOWN_BP", "DOWN_BVRP", False),
    RegionRunSpec("Total_region", "rp_global_sum", "TOTAL", "TOTAL_BP", "TOTAL_BVRP", False),
)


def _linreg_change(
    y: pd.Series,
    x: pd.Series,
    *,
    min_n: int = 8,
) -> dict | None:
    m = y.notna() & x.notna() & np.isfinite(y.astype(float)) & np.isfinite(x.astype(float))
    yy = y.loc[m].astype(float).to_numpy()
    xx = x.loc[m].astype(float).to_numpy()
    if len(xx) < min_n:
        return None
    try:
        from scipy.stats import linregress
    except ImportError:
        return {
            "n": len(xx),
            "intercept": float(np.nan),
            "slope": float(np.nan),
            "se_slope": float(np.nan),
            "t statistics": float(np.nan),
            "p value": float(np.nan),
            "r2": float(np.nan),
            "note": "need scipy for regression",
        }

    res = linregress(xx, yy)
    se = getattr(res, "stderr", float("nan")) or float("nan")
    t_sl = float(res.slope / se) if se and not np.isnan(se) and se > 0 else float("nan")
    return {
        "n": len(xx),
        "intercept": float(res.intercept),
        "slope": float(res.slope),
        "se_slope": float(se),
        "t statistics": t_sl,
        "p value": float(res.pvalue),
        "r2": float(res.rvalue**2),
        "note": "",
    }


def _prepare_merged_changes(
    merged: pd.DataFrame,
    vcol: str,
    *,
    negate_bvrp_for_display: bool,
) -> tuple[pd.Series, pd.Series, pd.Series]:
    """对齐于同一日历行的 forward ΔBP、forward ΔBVRP（经济符号）、ΔX。"""
    m = merged.sort_values("Date").drop_duplicates(subset=["Date"], keep="first")
    bp = m["BP"].astype(float)
    raw_bv = m["BVRP"].astype(float)
    bvrp_ec = (-raw_bv) if negate_bvrp_for_display else raw_bv
    x = m[vcol].astype(float)

    d_bp_fwd = bp.shift(-1) - bp
    d_bvrp_fwd = bvrp_ec.shift(-1) - bvrp_ec
    d_x = x.diff()
    return d_bp_fwd, d_bvrp_fwd, d_x


def run_regressions_for_observable(
    obs_path: Path,
    merged: pd.DataFrame,
    vcol: str,
    run: BoundsRunConfig,
    rspec: RegionRunSpec,
) -> list[dict]:
    lab = observable_plot_label(obs_path, vcol)
    d_bp, d_bv, d_x = _prepare_merged_changes(
        merged,
        vcol,
        negate_bvrp_for_display=run.negate_bvrp_for_display,
    )
    out: list[dict] = []
    base = {
        "bounds_key": run.key,
        "region": rspec.region_code,
        "observable_file": obs_path.name,
        "observable_column": vcol,
        "observable_label": lab,
    }
    for dep_name, yser in ((rspec.dep_bp, d_bp), (rspec.dep_bvrp, d_bv)):
        if rr := _linreg_change(yser, d_x):
            out.append(
                {
                    **base,
                    "dependent": dep_name,
                    **{
                        k: rr[k]
                        for k in (
                            "n",
                            "intercept",
                            "slope",
                            "se_slope",
                            "t statistics",
                            "p value",
                            "r2",
                            "note",
                        )
                    },
                }
            )
    return out


def _build_summary_dataframe(rows: list[dict]) -> pd.DataFrame:
    """按 p value 升序排列，便于一眼看到最显著的系数。"""
    df = pd.DataFrame(rows).copy()
    df["stars"] = df["p value"].map(_sig_stars)
    df["sig_level"] = df["p value"].map(_sig_level)
    df = df.sort_values("p value", na_position="last", kind="mergesort")
    return df[list(SUMMARY_COLUMNS)]


def _export_summary_xlsx(out_path: Path, summary: pd.DataFrame) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return False

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "all_by_p"

    sig10 = summary[summary["p value"].notna() & (summary["p value"] < 0.10)].copy()
    sig5 = summary[summary["p value"].notna() & (summary["p value"] < 0.05)].copy()

    header_font = Font(bold=True, name="Times New Roman", size=10)
    body_font = Font(name="Times New Roman", size=10)

    def write_sheet(ws, frame: pd.DataFrame) -> None:
        if frame.empty:
            ws.cell(1, 1, "(empty)")
            return
        for j, col in enumerate(frame.columns, start=1):
            c = ws.cell(1, j, col)
            c.font = header_font
        for i, rec in enumerate(frame.itertuples(index=False), start=2):
            for j, val in enumerate(rec, start=1):
                ws.cell(i, j, val).font = body_font

    write_sheet(ws0, summary)
    ws1 = wb.create_sheet("p_lt_10pct")
    write_sheet(ws1, sig10)
    ws2 = wb.create_sheet("p_lt_5pct")
    write_sheet(ws2, sig5)

    note = wb.create_sheet("note")
    note.cell(1, 1, "Stars on slope (two-sided p): *** p<0.01, ** p<0.05, * p<0.10.")
    note.cell(1, 1).font = body_font

    wb.save(out_path)
    return True


def export_run_results(
    run: BoundsRunConfig,
    rows: list[dict],
    rspec: RegionRunSpec,
) -> tuple[Path | None, Path | None]:
    """写出 <file_prefix>_delta_on_observable_delta.csv 与 _regression_summary.csv；UP 可另写 xlsx。"""
    if not rows:
        return None, None
    out_dir = OUTPUT_ROOT / run.output_subdir
    out_dir.mkdir(parents=True, exist_ok=True)
    pre = rspec.file_prefix
    detail_path = out_dir / f"{pre}_delta_on_observable_delta.csv"
    pd.DataFrame(rows)[list(RESULT_COLUMNS)].to_csv(detail_path, index=False)

    summary = _build_summary_dataframe(rows)
    summary_csv = out_dir / f"{pre}_regression_summary.csv"
    summary.to_csv(summary_csv, index=False)

    if not rspec.write_summary_xlsx:
        return detail_path, None

    xlsx_path = out_dir / f"{pre}_regression_summary.xlsx"
    if _export_summary_xlsx(xlsx_path, summary):
        return detail_path, xlsx_path
    return detail_path, None


def main() -> None:
    os.chdir(BASE_DIR)
    ensure_derived_observables()
    obs_files = list_observable_files()
    if not obs_files:
        raise FileNotFoundError("未找到可观测变量文件（Data/DATA Observables）")
    if not REGRESSION_RUNS:
        raise RuntimeError("S14_correlation.BOUNDS_RUNS 中缺少 CL23 配置")

    for run in REGRESSION_RUNS:
        for rspec in CL23_REGION_SPECS:
            bounds = load_cl23_region_bounds(run, rspec.rp_excel_col)
            if bounds is None:
                print(
                    f"[失败 {rspec.region_code}] 未读取 CL23 列 {rspec.rp_excel_col}。"
                    f"请确认 {CL23_BP_BVRP_XLSX} 存在且表 BP/VRP 含 date 与 {rspec.rp_excel_col}。"
                )
                continue

            bp_r, bv_r = bounds
            all_rows: list[dict] = []
            for obs_path in obs_files:
                try:
                    obs, freq = load_observable(obs_path)
                except Exception as e:
                    print(f"[{run.key}][{rspec.region_code}][跳过] {obs_path.name}: {e}")
                    continue
                vcols = observable_value_columns(obs)
                if not vcols:
                    continue
                try:
                    merged = merge_for_plot(obs, freq, bp_r, bv_r)
                except Exception as e:
                    print(
                        f"[{run.key}][{rspec.region_code}][跳过] merge {obs_path.name}: {e}"
                    )
                    continue
                if merged.empty or len(merged) < 4:
                    continue
                for vcol in vcols:
                    all_rows.extend(
                        run_regressions_for_observable(
                            obs_path, merged, vcol, run, rspec
                        )
                    )

            detail_p, summary_xlsx = export_run_results(run, all_rows, rspec)
            if detail_p:
                summ_csv = detail_p.parent / f"{rspec.file_prefix}_regression_summary.csv"
                msg = (
                    f"[{run.key}][{rspec.region_code}] 明细 {detail_p}；汇总 {summ_csv} "
                    f"（按 p 升序，共 {len(all_rows)} 条）"
                )
                if summary_xlsx:
                    msg += f"；Excel {summary_xlsx}"
                elif rspec.write_summary_xlsx:
                    msg += "（未写 .xlsx：可 pip install openpyxl）"
                print(msg)
            else:
                print(
                    f"[{run.key}][{rspec.region_code}] 无有效回归结果（检查对齐后样本长度）"
                )


if __name__ == "__main__":
    main()
