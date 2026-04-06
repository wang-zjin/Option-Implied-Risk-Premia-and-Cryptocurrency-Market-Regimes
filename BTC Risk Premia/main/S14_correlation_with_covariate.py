"""
外生变量 vs 估计得到的 BP / BVRP：对齐、作图、IC 表。

默认跑三套界：CL20、CL23、SS25。
- CL20：Data/RiskPremia/BP_CL20.csv 与 BVRP_CL20.csv（Date；BP 或 Lower_Bound；BVRP 或 Upper_Bound）。默认仍按 UBU2 对 BVRP 作 − 再画/算 IC（见 negate_bvrp_for_display）。
- CL23：Data/RiskPremia/ 下 **BP_BVRP_CL23.xlsx**（优先）或 **BP_VRP_CL23.xlsx**，子表 BP + VRP；SS25：BP_BVRP_SS25_ttm27.csv 单表。
- 产出：S14_correlation/<CL20|CL23|SS25>/ 下 PNG 与 S14_IC_correlation.xlsx。

脚本结构：配置 → 显示名 → 衍生数据 → 合并 → Observables → 作图 → IC → main。
"""

from __future__ import annotations

import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.lines import Line2D

# -----------------------------------------------------------------------------
# Matplotlib（英文环境）
# -----------------------------------------------------------------------------
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["font.sans-serif"] = [
    "DejaVu Sans",
    "Arial",
    "Helvetica",
    "Liberation Sans",
]

# =============================================================================
# 路径与开关
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent
OBSERVABLES_DIR = BASE_DIR / "Data" / "DATA Observables"
S14_CORR_ROOT = BASE_DIR / "S14_correlation"
EXPORT_IC_XLSX = True

RISK_PREMIA_DIR = BASE_DIR / "Data" / "RiskPremia"
BP_CL20_CSV = RISK_PREMIA_DIR / "BP_CL20.csv"
BVRP_CL20_CSV = RISK_PREMIA_DIR / "BVRP_CL20.csv"


def resolve_cl23_combined_xlsx_path() -> Path:
    """CL23 工作簿：优先 BP_BVRP_CL23.xlsx，否则 BP_VRP_CL23.xlsx（同一结构：子表 BP、VRP）。"""
    primary = RISK_PREMIA_DIR / "BP_BVRP_CL23.xlsx"
    alt = RISK_PREMIA_DIR / "BP_VRP_CL23.xlsx"
    if primary.is_file():
        return primary
    if alt.is_file():
        return alt
    return primary


CL23_BP_BVRP_XLSX = resolve_cl23_combined_xlsx_path()
SS25_BP_BVRP_CSV = RISK_PREMIA_DIR / "BP_BVRP_SS25_ttm27.csv"


@dataclass(frozen=True)
class BoundsRunConfig:
    """一套 BP/BVRP 与输出目录 S14_correlation/<output_subdir>/。

    输入三选一：
    (1) combined_bp_bvrp_path 单表 .csv/.xlsx（首张表），列见 load_combined_bp_bvrp_table；
    (2) combined_bp_bvrp_path 为 xlsx 且指定 combined_xlsx_bp_sheet + combined_xlsx_vrp_sheet 两子表；
    (3) bp_csv + bvrp_csv 两 CSV（见 load_bounds_series）。
    negate_bvrp_for_display：True 时作图与 IC 的 BVRP 为 −BVRP 列（UBU2 约定）；False 时直接用 BVRP 列（SS25 的 VRP）。
    """

    key: str
    output_subdir: str
    legend_tag: str
    ic_bounds_note: str
    negate_bvrp_for_display: bool
    ic_bvrp_footnote: str
    combined_bp_bvrp_path: Path | None = None
    # 若 xlsx 中 BP、方差溢价分属两个子表（如 sheet「BP」「VRP」），填二者名称
    combined_xlsx_bp_sheet: str | None = None
    combined_xlsx_vrp_sheet: str | None = None
    bp_csv: Path | None = None
    bvrp_csv: Path | None = None

    @property
    def output_dir(self) -> Path:
        return S14_CORR_ROOT / self.output_subdir

    @property
    def ic_xlsx_path(self) -> Path:
        return self.output_dir / "S14_IC_correlation.xlsx"


BOUNDS_RUNS: tuple[BoundsRunConfig, ...] = (
    BoundsRunConfig(
        key="CL20",
        output_subdir="CL20",
        legend_tag="CL20",
        ic_bounds_note="CL20 (Chabi-Yo & Loudis 2020)",
        negate_bvrp_for_display=True,
        ic_bvrp_footnote=(
            "BVRP from BVRP_CL20.csv: default −UBU2 if values are raw Upper_Bound; "
            "if already economic BVRP, set negate_bvrp_for_display=False in BOUNDS_RUNS."
        ),
        bp_csv=BP_CL20_CSV,
        bvrp_csv=BVRP_CL20_CSV,
    ),
    BoundsRunConfig(
        key="CL23",
        output_subdir="CL23",
        legend_tag="CL23",
        ic_bounds_note="CL23 (Chabi-Yo & Loudis, 2023)",
        negate_bvrp_for_display=False,
        ic_bvrp_footnote=(
            "BVRP from CL23 workbook (BP_BVRP_CL23.xlsx or BP_VRP_CL23.xlsx) sheet VRP (no −UBU2); "
            "if values are raw UBU2, set negate_bvrp_for_display=True."
        ),
        combined_bp_bvrp_path=CL23_BP_BVRP_XLSX,
        combined_xlsx_bp_sheet="BP",
        combined_xlsx_vrp_sheet="VRP",
    ),
    BoundsRunConfig(
        key="SS25",
        output_subdir="SS25",
        legend_tag="SS25",
        ic_bounds_note="SS25 (Schreindorfer & Sichert 2025), τ=27d",
        negate_bvrp_for_display=False,
        ic_bvrp_footnote=(
            "BVRP row uses column VRP from BP_BVRP_SS25_ttm27.csv (no −UBU2 transform)."
        ),
        combined_bp_bvrp_path=SS25_BP_BVRP_CSV,
    ),
)

TRADING_ACTIVITY_SOURCE = (
    BASE_DIR / "Data/processed/20172022_1_3_6_duplicates_moneyness_tau.csv"
)
TRADING_ACTIVITY_OUT = OBSERVABLES_DIR / "BTC_options_trading_activity.csv"
TRADING_ACTIVITY_CHUNK = 400_000
TRADING_ACTIVITY_FORCE_REBUILD = False

MOMENTUM_BTC_PRICE_CSV = BASE_DIR / "Data" / "BTC_USD_Quandl_2011_2023.csv"
MOMENTUM_OUT = OBSERVABLES_DIR / "BTC_momentum_weekly_7d.csv"
MOMENTUM_TRADING_DAYS = 7
MOMENTUM_FORCE_REBUILD = False

# -----------------------------------------------------------------------------
# 图中显示名（Sentix 等：文件名 / 列名 → 统一标签）
# -----------------------------------------------------------------------------
OBSERVABLE_DISPLAY_BY_STEM: dict[str, str] = {
    "sentiment headline index": "Sentiment Headline Index",
    "overconfidence index": "Overconfidence Index",
    "sentiment institutional investors": "Sentiment Institutional Investors",
    "sentiment individual investors": "Sentiment Individual Investors",
    "sentiment bullish %": "Sentiment Bullish",
    "sentiment bearish %": "Sentiment Bearish",
}

OBSERVABLE_COLUMN_DISPLAY: dict[tuple[str, str], str] = {
    ("btc_options_trading_activity.csv", "total_volume"): "Trading volume",
    ("btc_options_trading_activity.csv", "value_weighted_volume"): (
        "Value-weighted volume"
    ),
    ("btc_options_trading_activity.csv", "put_call_ratio"): "Put-call ratio",
    ("btc_momentum_weekly_7d.csv", "momentum_avg_return"): (
        "Momentum (7d avg return)"
    ),
}

IC_NUMERIC_HEADERS: tuple[str, ...] = (
    "label",
    "source_file",
    "column",
    "n",
    "pearson_bp_r",
    "pearson_bp_p",
    "pearson_bvrp_r",
    "pearson_bvrp_p",
    "spearman_bp_r",
    "spearman_bp_p",
    "spearman_bvrp_r",
    "spearman_bvrp_p",
)


def observable_display_from_stem(stem: str) -> str | None:
    return OBSERVABLE_DISPLAY_BY_STEM.get(stem.strip().lower())


def observable_column_display(path: Path, vcol: str) -> str | None:
    return OBSERVABLE_COLUMN_DISPLAY.get((path.name.lower(), vcol))


def observable_plot_label(obs_path: Path, vcol: str) -> str:
    """与作图一致的系列名（总标题 / 图例）。"""
    stem_l = obs_path.stem.lower()
    if stem_l == "crix_data" and vcol == "price":
        return "CRIX"
    if stem_l == "crix_monthly" and vcol == "CRIX_approx":
        return "CRIX (monthly)"
    if stem_l == "google trends bitcoin monthly" and vcol == "bitcoin":
        return "Google trend: BTC"
    if (cn := observable_column_display(obs_path, vcol)) is not None:
        return cn
    if (named := observable_display_from_stem(obs_path.stem)) is not None:
        return named
    return vcol


# =============================================================================
# 写入 Observables 的衍生序列
# =============================================================================
def build_btc_options_trading_activity_csv() -> None:
    """
    日度：total_volume = Σ volume（C+P）；value_weighted_volume = Σ volume_optionprice；
    put_call_ratio = P/(P+C)（P、C 为 put/call 的 volume 日度和）。
    """
    if not TRADING_ACTIVITY_SOURCE.is_file():
        print(f"[提示] 未找到 {TRADING_ACTIVITY_SOURCE}，跳过 Trading activity。")
        return
    if TRADING_ACTIVITY_OUT.is_file() and not TRADING_ACTIVITY_FORCE_REBUILD:
        return

    chunks: list[pd.DataFrame] = []
    for ch in pd.read_csv(
        TRADING_ACTIVITY_SOURCE,
        chunksize=TRADING_ACTIVITY_CHUNK,
        usecols=["date", "putcall", "volume", "volume_optionprice"],
    ):
        ch = ch[ch["putcall"].isin(["C", "P"])]
        put_sum = (
            ch.loc[ch["putcall"] == "P"]
            .groupby("date", as_index=False)
            .agg(
                put_volume=("volume", "sum"),
                put_vw=("volume_optionprice", "sum"),
            )
        )
        call_sum = (
            ch.loc[ch["putcall"] == "C"]
            .groupby("date", as_index=False)
            .agg(
                call_volume=("volume", "sum"),
                call_vw=("volume_optionprice", "sum"),
            )
        )
        g = put_sum.merge(call_sum, on="date", how="outer").fillna(0.0)
        g["value_weighted_volume"] = g["put_vw"] + g["call_vw"]
        g = g[["date", "put_volume", "call_volume", "value_weighted_volume"]]
        chunks.append(g)

    out = pd.concat(chunks, ignore_index=True)
    out = out.groupby("date", as_index=False).sum(numeric_only=True)
    out["total_volume"] = out["put_volume"] + out["call_volume"]
    den = out["total_volume"]
    out["put_call_ratio"] = (out["put_volume"] / den).where(den > 0)
    out = out.rename(columns={"date": "Date"})
    out["Date"] = pd.to_datetime(out["Date"])
    out = (
        out[["Date", "total_volume", "value_weighted_volume", "put_call_ratio"]]
        .sort_values("Date")
        .reset_index(drop=True)
    )
    OBSERVABLES_DIR.mkdir(parents=True, exist_ok=True)
    out.to_csv(TRADING_ACTIVITY_OUT, index=False)
    print(f"已生成 {TRADING_ACTIVITY_OUT}（{len(out)} 个交易日）。")


def build_btc_momentum_weekly_csv() -> None:
    """
    mom_t = mean(r_{t-1},…,r_{t-K})，r 为 Adj.Close 日收益；shift(1) 便于与当日 CL20 对齐。
    """
    if not MOMENTUM_BTC_PRICE_CSV.is_file():
        print(f"[提示] 未找到 {MOMENTUM_BTC_PRICE_CSV}，跳过 Momentum。")
        return
    if MOMENTUM_OUT.is_file() and not MOMENTUM_FORCE_REBUILD:
        return

    df = pd.read_csv(MOMENTUM_BTC_PRICE_CSV, parse_dates=["Date"])
    df = df.sort_values("Date").reset_index(drop=True)
    price_col = "Adj.Close"
    if price_col not in df.columns:
        raise ValueError(f"{MOMENTUM_BTC_PRICE_CSV} 缺少列 {price_col}")

    r = df[price_col].pct_change()
    k = MOMENTUM_TRADING_DAYS
    out = pd.DataFrame(
        {
            "Date": df["Date"],
            "momentum_avg_return": r.shift(1).rolling(k, min_periods=k).mean(),
        }
    ).dropna(subset=["momentum_avg_return"])
    out = out.reset_index(drop=True)
    OBSERVABLES_DIR.mkdir(parents=True, exist_ok=True)
    out.to_csv(MOMENTUM_OUT, index=False)
    print(f"已生成 {MOMENTUM_OUT}（{len(out)} 日，K={k}）。")


def ensure_derived_observables() -> None:
    build_btc_options_trading_activity_csv()
    build_btc_momentum_weekly_csv()


# =============================================================================
# CL20 与对齐
# =============================================================================
def load_bounds_series(bp_csv: Path, bvrp_csv: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """两表 CSV：Date；BP 或 Lower_Bound；BVRP 或 Upper_Bound（读入后统一为 BP / BVRP 列）。"""
    bp_raw = pd.read_csv(bp_csv)
    bv_raw = pd.read_csv(bvrp_csv)
    bp_raw.columns = [str(c).strip() for c in bp_raw.columns]
    bv_raw.columns = [str(c).strip() for c in bv_raw.columns]
    if "Date" not in bp_raw.columns or "Date" not in bv_raw.columns:
        raise ValueError("BP/BVRP CSV 均需要 Date 列")
    if "BP" in bp_raw.columns:
        bp_vals = pd.to_numeric(bp_raw["BP"], errors="coerce")
    elif "Lower_Bound" in bp_raw.columns:
        bp_vals = pd.to_numeric(bp_raw["Lower_Bound"], errors="coerce")
    else:
        raise ValueError(f"{bp_csv.name} 需要 BP 或 Lower_Bound 列")
    if "BVRP" in bv_raw.columns:
        bv_vals = pd.to_numeric(bv_raw["BVRP"], errors="coerce")
    elif "Upper_Bound" in bv_raw.columns:
        bv_vals = pd.to_numeric(bv_raw["Upper_Bound"], errors="coerce")
    else:
        raise ValueError(f"{bvrp_csv.name} 需要 BVRP 或 Upper_Bound 列")
    bp_dt = _parse_yyyymmdd_or_datetime(bp_raw["Date"])
    bv_dt = _parse_yyyymmdd_or_datetime(bv_raw["Date"])
    bp = pd.DataFrame({"Date": bp_dt, "BP": bp_vals}).dropna(subset=["Date"])
    bvrp = pd.DataFrame({"Date": bv_dt, "BVRP": bv_vals}).dropna(subset=["Date"])
    return bp, bvrp


def _parse_yyyymmdd_or_datetime(s: pd.Series) -> pd.Series:
    """SS25 表常见 Date=20170905；CL20 为可解析日期字符串。"""
    if pd.api.types.is_datetime64_any_dtype(s):
        return pd.to_datetime(s, errors="coerce")
    str_s = s.astype(str).str.strip()
    parsed = pd.to_datetime(str_s, format="%Y%m%d", errors="coerce")
    if parsed.notna().sum() >= max(3, len(s) // 2):
        return parsed
    return pd.to_datetime(s, errors="coerce")


def _read_bounds_table(path: Path) -> pd.DataFrame:
    suf = path.suffix.lower()
    if suf == ".xlsx":
        return pd.read_excel(path, sheet_name=0, engine="openpyxl")
    if suf == ".xls":
        return pd.read_excel(path, sheet_name=0)
    return pd.read_csv(path)


def _normalize_combined_table_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip().replace("\ufeff", "") for c in out.columns]
    cl = {c.lower().replace(" ", "_"): c for c in out.columns}
    if "Date" not in out.columns and "date" in cl:
        out = out.rename(columns={cl["date"]: "Date"})
    return out


def _column_key_index(df: pd.DataFrame) -> dict[str, str]:
    """规范化列键（小写、空格→_）→ 实际列名，便于识别 bp / Lower Bound 等。"""
    idx: dict[str, str] = {}
    for c in df.columns:
        raw = str(c).strip().replace("\ufeff", "")
        k = raw.lower().replace(" ", "_").replace("-", "_")
        if k not in idx:
            idx[k] = raw
    return idx


def _infer_bp_and_variance_columns(df: pd.DataFrame) -> tuple[str, str]:
    """返回 (bp 列名, 方差溢价列名)；方差列读入后统一为内部 BVRP。"""
    idx = _column_key_index(df)
    bp_keys = ("bp", "lower_bound", "lbu", "rp_global_sum")
    var_keys = (
        "vrp",
        "bvrp",
        "upper_bound",
        "ubu2",
        "ub2",
        "vrp_global",
        "bvrp_global",
        "vrp_global_sum",
    )
    bp_col = next((idx[k] for k in bp_keys if k in idx), None)
    var_col = next((idx[k] for k in var_keys if k in idx), None)
    if bp_col is None or var_col is None:
        raise ValueError(
            "单表需要 BP（或 Lower_Bound/LBU）与 VRP/BVRP（或 Upper_Bound/UBU2）；"
            f"当前列: {list(df.columns)}"
        )
    return bp_col, var_col


def load_combined_bp_bvrp_table(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """单表 .csv/.xlsx：Date；BP 或 Lower_Bound；VRP / BVRP / Upper_Bound → 内部列 BVRP。"""
    df = _normalize_combined_table_columns(_read_bounds_table(path))
    idx = _column_key_index(df)
    date_col = idx.get("date") or idx.get("datum")
    if date_col is None and "Date" in df.columns:
        date_col = "Date"
    if date_col is None:
        raise ValueError(
            f"{path.name} 需要日期列（Date）；当前列: {list(df.columns)}"
        )
    bp_col, var_col = _infer_bp_and_variance_columns(df)
    dt = _parse_yyyymmdd_or_datetime(df[date_col])
    one = pd.DataFrame(
        {
            "Date": dt,
            "BP": pd.to_numeric(df[bp_col], errors="coerce"),
            "BVRP": pd.to_numeric(df[var_col], errors="coerce"),
        }
    ).dropna(subset=["Date"])
    return one[["Date", "BP"]], one[["Date", "BVRP"]]


def _date_column_name(df: pd.DataFrame) -> str:
    idx = _column_key_index(df)
    if "date" in idx:
        return idx["date"]
    if "Date" in df.columns:
        return "Date"
    raise ValueError(f"需要日期列 Date；当前列: {list(df.columns)}")


def _non_date_value_column(df: pd.DataFrame, date_col: str) -> str:
    """除日期外唯一（或优先命名的）数值列，用于分表 BP/VRP 各只有一列的情况。"""
    others = [c for c in df.columns if c != date_col]
    if not others:
        raise ValueError(f"除 {date_col} 外需要至少一列数值；当前列: {list(df.columns)}")
    idx = _column_key_index(df)
    for key in ("rp_global_sum", "bp", "vrp", "bvrp", "lower_bound", "upper_bound"):
        if key in idx and idx[key] != date_col:
            return idx[key]
    return others[0]


def _read_excel_sheet(path: Path, sheet: str) -> pd.DataFrame:
    if path.suffix.lower() == ".xlsx":
        return pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    return pd.read_excel(path, sheet_name=sheet)


def load_combined_excel_bp_vrp_sheets(
    path: Path,
    bp_sheet: str,
    vrp_sheet: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """同一 xlsx 内两个子表：各含日期 + 一列数值，分别作为 BP 与 BVRP。"""
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError("双子表模式仅适用于 .xlsx/.xls")
    bp_df = _normalize_combined_table_columns(_read_excel_sheet(path, bp_sheet))
    v_df = _normalize_combined_table_columns(_read_excel_sheet(path, vrp_sheet))
    d_bp = _date_column_name(bp_df)
    d_v = _date_column_name(v_df)
    v_bp = _non_date_value_column(bp_df, d_bp)
    v_v = _non_date_value_column(v_df, d_v)
    bp = pd.DataFrame(
        {
            "Date": _parse_yyyymmdd_or_datetime(bp_df[d_bp]),
            "BP": pd.to_numeric(bp_df[v_bp], errors="coerce"),
        }
    ).dropna(subset=["Date"])
    bvrp = pd.DataFrame(
        {
            "Date": _parse_yyyymmdd_or_datetime(v_df[d_v]),
            "BVRP": pd.to_numeric(v_df[v_v], errors="coerce"),
        }
    ).dropna(subset=["Date"])
    return bp, bvrp


def load_bounds_for_run(run: BoundsRunConfig) -> tuple[pd.DataFrame, pd.DataFrame]:
    if run.combined_bp_bvrp_path is not None:
        if run.combined_xlsx_bp_sheet and run.combined_xlsx_vrp_sheet:
            return load_combined_excel_bp_vrp_sheets(
                run.combined_bp_bvrp_path,
                run.combined_xlsx_bp_sheet,
                run.combined_xlsx_vrp_sheet,
            )
        return load_combined_bp_bvrp_table(run.combined_bp_bvrp_path)
    if run.bp_csv is not None and run.bvrp_csv is not None:
        return load_bounds_series(run.bp_csv, run.bvrp_csv)
    raise ValueError(f"BOUNDS_RUNS 配置缺少输入路径: {run.key}")


def bounds_run_inputs_ready(run: BoundsRunConfig) -> bool:
    if run.combined_bp_bvrp_path is not None:
        return run.combined_bp_bvrp_path.is_file()
    if run.bp_csv is not None and run.bvrp_csv is not None:
        return run.bp_csv.is_file() and run.bvrp_csv.is_file()
    return False


def _monthly_mean_cl20(bp: pd.DataFrame, bvrp: pd.DataFrame) -> pd.DataFrame:
    m = bp.merge(bvrp, on="Date", how="inner").set_index("Date").sort_index()
    g = m.groupby(pd.Grouper(freq="ME")).mean(numeric_only=True).reset_index()
    return g.rename(columns={"Date": "month_end"})


def merge_for_plot(
    obs: pd.DataFrame,
    freq: str,
    bp: pd.DataFrame,
    bvrp: pd.DataFrame,
) -> pd.DataFrame:
    val_cols = [c for c in obs.columns if c != "Date"]
    if not val_cols:
        raise ValueError("可观测数据没有数值列")

    if freq == "M":
        obs_m = obs.copy()
        obs_m["month"] = obs_m["Date"].dt.to_period("M")
        cl20_m = _monthly_mean_cl20(bp, bvrp)
        cl20_m["month"] = cl20_m["month_end"].dt.to_period("M")
        m = obs_m.merge(cl20_m[["month", "BP", "BVRP"]], on="month", how="inner")
        return m.drop(columns=["month"])

    cl20_d = bp.merge(bvrp, on="Date", how="inner")
    return obs.merge(cl20_d, on="Date", how="inner")


def _infer_freq_from_median_gap(dates: pd.Series) -> Literal["D", "M"]:
    gap = dates.sort_values().diff().median()
    return "M" if gap is not pd.NaT and gap.days > 20 else "D"


# =============================================================================
# Observables：列举与读取
# =============================================================================
def list_observable_files() -> list[Path]:
    return sorted(
        [*OBSERVABLES_DIR.glob("*.csv")]
        + [*OBSERVABLES_DIR.glob("*.xlsx")]
        + [*OBSERVABLES_DIR.glob("*.xls")]
    )


def observable_value_columns(obs: pd.DataFrame) -> list[str]:
    return [
        c
        for c in obs.columns
        if c != "Date" and not _skip_observable_column(c)
    ]


def _skip_observable_column(name: str) -> bool:
    n = str(name).strip()
    if not n:
        return True
    nl = n.lower()
    if nl in ("code", "id", "index", "key"):
        return True
    return nl.startswith("unnamed")


def _load_observable_excel(path: Path) -> tuple[pd.DataFrame, Literal["D", "M"]]:
    suf = path.suffix.lower()
    try:
        if suf == ".xlsx":
            df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
        else:
            df = pd.read_excel(path, sheet_name=0)
    except ImportError as e:
        raise RuntimeError("读取 .xlsx 请先安装: pip install openpyxl") from e
    except Exception as e:
        raise RuntimeError(f"无法读取 Excel {path.name}: {e}") from e

    df.columns = [str(c).strip() for c in df.columns]
    cl = {c.lower(): c for c in df.columns}

    if "year" in cl and "month" in cl and "date" not in cl:
        y = pd.to_numeric(df[cl["year"]], errors="coerce")
        mo = pd.to_numeric(df[cl["month"]], errors="coerce")
        ok = y.notna() & mo.notna()
        df = df.loc[ok].copy()
        y, mo = y.loc[ok].astype(int), mo.loc[ok].astype(int)
        out = pd.DataFrame({"Date": pd.to_datetime(dict(year=y, month=mo, day=1))})
        for c in df.columns:
            lc = str(c).lower()
            if lc in ("year", "month") or _skip_observable_column(c):
                continue
            out[c] = pd.to_numeric(df[c], errors="coerce")
        return out, "M"

    date_col = None
    for key in ("date", "time", "datetime", "week"):
        if key in cl:
            date_col = cl[key]
            break
    if date_col is None and "month" in cl:
        ttry = pd.to_datetime(df[cl["month"]], errors="coerce", dayfirst=True)
        if ttry.notna().sum() >= max(3, len(df) // 3):
            date_col = cl["month"]
    if date_col is None:
        c0 = df.columns[0]
        ttry = pd.to_datetime(df[c0], errors="coerce", dayfirst=True)
        if ttry.notna().sum() >= max(3, len(df) // 3):
            date_col = c0
    if date_col is None:
        raise ValueError("未识别日期列：Date/Time、Year+Month 或首列为日期")

    out = pd.DataFrame(
        {"Date": pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)}
    )
    for c in df.columns:
        if c == date_col or _skip_observable_column(c):
            continue
        out[c] = pd.to_numeric(df[c], errors="coerce")
    out = out.dropna(subset=["Date"])
    if len(out) < 2:
        return out, "D"
    return out, _infer_freq_from_median_gap(out["Date"])


def load_observable(path: Path) -> tuple[pd.DataFrame, Literal["D", "M"]]:
    name = path.name.lower()
    if name.endswith((".xlsx", ".xls")):
        return _load_observable_excel(path)

    if name == "global_policy_uncertainty_data.csv":
        df = pd.read_csv(path)
        y = pd.to_numeric(df["Year"], errors="coerce")
        mo = pd.to_numeric(df["Month"], errors="coerce")
        ok = y.notna() & mo.notna()
        df = df.loc[ok].copy()
        y, mo = y.loc[ok].astype(int), mo.loc[ok].astype(int)
        out = pd.DataFrame({"Date": pd.to_datetime(dict(year=y, month=mo, day=1))})
        for c in df.columns:
            if c not in ("Year", "Month"):
                out[c] = pd.to_numeric(df[c], errors="coerce")
        return out, "M"

    if name == "google trends bitcoin monthly.csv":
        df = pd.read_csv(path, quotechar='"')
        df.columns = [c.strip().strip('"') for c in df.columns]
        time_col = "Time" if "Time" in df.columns else df.columns[0]
        out = pd.DataFrame(
            {"Date": pd.to_datetime(df[time_col].astype(str).str.strip('"'))}
        )
        for c in df.columns:
            if c != time_col:
                out[c] = pd.to_numeric(df[c], errors="coerce")
        return out, "M"

    if name == "crix_monthly.csv":
        return pd.read_csv(path, parse_dates=["Date"]), "M"

    if name == "ads_index_most_current_vintage.csv":
        df = pd.read_csv(path)
        raw = df["Date"].astype(str).str.replace(":", "-", regex=False)
        out = pd.DataFrame({"Date": pd.to_datetime(raw, errors="coerce")})
        for c in df.columns:
            if c == "Date":
                continue
            if pd.api.types.is_numeric_dtype(df[c]) or df[c].dtype == object:
                out[c] = pd.to_numeric(df[c], errors="coerce")
        return out.dropna(subset=["Date"]), "D"

    if name == "crix_data.csv":
        df = pd.read_csv(path, parse_dates=["date"])
        return df.rename(columns={"date": "Date"}), "D"

    df = pd.read_csv(path)
    first = df.columns[0]
    try:
        dates = pd.to_datetime(df[first], errors="coerce")
    except Exception:
        dates = pd.to_datetime(df[first].astype(str), errors="coerce")
    out = pd.DataFrame({"Date": dates})
    for c in df.columns[1:]:
        out[c] = pd.to_numeric(df[c], errors="coerce")
    out = out.dropna(subset=["Date"])
    return out, _infer_freq_from_median_gap(out["Date"])


# =============================================================================
# 文件名桩（输出 PNG）
# =============================================================================
def safe_stub(path: Path) -> str:
    s = re.sub(r"[^\w\-]+", "_", path.stem, flags=re.UNICODE).strip("_")
    return s or "observable"


def safe_col_stub(name: str) -> str:
    s = re.sub(r"[^\w\-]+", "_", str(name), flags=re.UNICODE).strip("_")
    return s or "col"


# =============================================================================
# 作图（仅对 BP/BVRP 插值，避免断线）
# =============================================================================
def interpolate_bounds_for_plot(merged: pd.DataFrame, date_col: str) -> pd.DataFrame:
    out = merged.sort_values(date_col).copy()
    if len(out) < 2:
        return out
    if out[date_col].duplicated().any():
        out = out.drop_duplicates(subset=[date_col], keep="first")
    dt = pd.to_datetime(out[date_col], errors="coerce")
    body = out.drop(columns=[date_col]).set_index(dt)
    for col in ("BP", "BVRP"):
        if col not in body.columns:
            continue
        s = pd.to_numeric(body[col], errors="coerce")
        s = s.interpolate(method="time", limit_direction="both")
        if s.isna().any():
            s = s.interpolate(method="index", limit_direction="both")
        body[col] = s
    restored = body.reset_index()
    return restored.rename(columns={restored.columns[0]: date_col})


def _bounds_legend_handles(obs_lab: str, legend_tag: str) -> list[Line2D]:
    return [
        Line2D(
            [0],
            [0],
            color="tab:blue",
            linewidth=1.2,
            linestyle="-",
            label=f"BP ({legend_tag})",
        ),
        Line2D(
            [0],
            [0],
            color="tab:red",
            linewidth=1.2,
            linestyle="-",
            label=f"BVRP ({legend_tag})",
        ),
        Line2D(
            [0],
            [0],
            color="tab:orange",
            linewidth=1.0,
            linestyle="--",
            label=obs_lab,
        ),
    ]


def plot_observable_vs_bounds(
    merged: pd.DataFrame,
    vcol: str,
    out_path: Path,
    *,
    legend_tag: str,
    negate_bvrp_for_display: bool,
    suptitle_label: str | None = None,
    observable_label: str | None = None,
) -> None:
    date_col = "Date" if "Date" in merged.columns else "month_end"
    merged = interpolate_bounds_for_plot(merged, date_col)
    t = merged[date_col]
    obs_lab = observable_label if observable_label is not None else vcol
    st_lab = suptitle_label if suptitle_label is not None else vcol

    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)

    ax0, ax0_r = axes[0], axes[0].twinx()
    ax0.plot(t, merged["BP"], color="tab:blue", linewidth=1.2)
    ax0_r.plot(t, merged[vcol], color="tab:orange", linewidth=1.0, linestyle="--")
    ax0.set_ylabel("BP")
    ax0_r.set_ylabel(obs_lab)
    ax0.grid(True, alpha=0.3)

    ax1, ax1_r = axes[1], axes[1].twinx()
    bvrp_series = merged["BVRP"].astype(float)
    bvrp_plot = (-bvrp_series) if negate_bvrp_for_display else bvrp_series
    ax1.plot(t, bvrp_plot, color="tab:red", linewidth=1.2)
    ax1_r.plot(t, merged[vcol], color="tab:orange", linewidth=1.0, linestyle="--")
    ax1.set_ylabel("BVRP")
    ax1_r.set_ylabel(obs_lab)
    ax1.grid(True, alpha=0.3)

    axes[-1].set_xlabel("Date")
    fig.suptitle(f"{st_lab} vs {legend_tag} BP and BVRP", fontsize=13)
    fig.tight_layout(rect=(0, 0.12, 1, 0.96))
    plt.setp(axes[-1].get_xticklabels(), rotation=0, ha="center")
    fig.legend(
        handles=_bounds_legend_handles(obs_lab, legend_tag),
        loc="lower center",
        bbox_to_anchor=(0.5, 0.08),
        ncol=3,
        fontsize=8,
        frameon=True,
    )
    fig.savefig(out_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def plot_all_observables_vs_bounds(
    files: list[Path],
    bp: pd.DataFrame,
    bvrp: pd.DataFrame,
    run: BoundsRunConfig,
) -> None:
    out_base = run.output_dir
    tag = run.key
    leg = run.legend_tag
    for obs_path in files:
        try:
            obs, freq = load_observable(obs_path)
        except Exception as e:
            print(f"[{tag}][跳过] {obs_path.name}: {e}")
            continue

        vcols = observable_value_columns(obs)
        if not vcols:
            print(f"[{tag}][跳过] {obs_path.name}: 无数值列")
            continue
        try:
            merged = merge_for_plot(obs, freq, bp, bvrp)
        except Exception as e:
            print(f"[{tag}][跳过] {obs_path.name} 合并失败: {e}")
            continue
        if merged.empty or len(merged) < 3:
            print(f"[{tag}][跳过] {obs_path.name}: 对齐后样本过少 ({len(merged)})")
            continue

        stub = safe_stub(obs_path)
        for vcol in vcols:
            if merged[vcol].notna().sum() < 3:
                print(f"[{tag}][跳过] {obs_path.name} 列 {vcol}: 有效点过少")
                continue
            out_png = (
                out_base
                / f"{stub}__{safe_col_stub(vcol)}_vs_{tag}_BP_BVRP.png"
            )
            lab = observable_plot_label(obs_path, vcol)
            plot_observable_vs_bounds(
                merged,
                vcol,
                out_png,
                legend_tag=leg,
                negate_bvrp_for_display=run.negate_bvrp_for_display,
                suptitle_label=lab,
                observable_label=lab,
            )
            print(f"[{tag}] 已保存: {out_png} (n={len(merged)}, freq={freq})")


# =============================================================================
# IC：相关 + 显著性星号 + xlsx
# =============================================================================
def _pairwise_corr(
    a: pd.Series,
    b: pd.Series,
    *,
    method: Literal["pearson", "spearman"],
) -> tuple[float, float, int]:
    m = a.notna() & b.notna()
    aa, bb = a[m].astype(float), b[m].astype(float)
    n = int(len(aa))
    if n < 3:
        return float("nan"), float("nan"), n
    if method == "pearson":
        try:
            from scipy.stats import pearsonr

            r, p = pearsonr(aa, bb)
            return float(r), float(p), n
        except ImportError:
            return float(aa.corr(bb)), float("nan"), n
    try:
        from scipy.stats import spearmanr

        r, p = spearmanr(aa, bb)
        return float(r), float(p), n
    except ImportError:
        return float(aa.corr(bb, method="spearman")), float("nan"), n


def _sig_stars(p: float) -> str:
    if math.isnan(p):
        return ""
    if p < 0.01:
        return "***"
    if p < 0.05:
        return "**"
    if p < 0.10:
        return "*"
    return ""


def _fmt_ic_cell(r: float, p: float) -> str:
    if math.isnan(r):
        return ""
    return f"{r:.2f}{_sig_stars(p)}"


def _make_unique_label(used: set[str], base: str, obs_path: Path) -> str:
    if base not in used:
        used.add(base)
        return base
    alt = f"{base} [{safe_stub(obs_path)}]"
    k, i = alt, 2
    while k in used:
        k = f"{alt}_{i}"
        i += 1
    used.add(k)
    return k


def _ic_row_for_column(
    obs_path: Path,
    merged: pd.DataFrame,
    vcol: str,
    used_labels: set[str],
    *,
    negate_bvrp_for_display: bool,
) -> dict | None:
    if merged[vcol].notna().sum() < 3:
        return None
    joint = (
        merged["BP"].notna() & merged["BVRP"].notna() & merged[vcol].notna()
    )
    if int(joint.sum()) < 3:
        return None
    sub = merged.loc[joint]
    label = _make_unique_label(
        used_labels, observable_plot_label(obs_path, vcol), obs_path
    )
    x = sub[vcol].astype(float)
    y_bp = sub["BP"].astype(float)
    raw_bvrp = sub["BVRP"].astype(float)
    bvrp_signed = (-raw_bvrp) if negate_bvrp_for_display else raw_bvrp

    r_pb, p_pb, n_pb = _pairwise_corr(y_bp, x, method="pearson")
    r_pv, p_pv, _ = _pairwise_corr(bvrp_signed, x, method="pearson")
    r_sb, p_sb, _ = _pairwise_corr(y_bp, x, method="spearman")
    r_sv, p_sv, _ = _pairwise_corr(bvrp_signed, x, method="spearman")

    return {
        "label": label,
        "source_file": obs_path.name,
        "column": vcol,
        "n": n_pb,
        "pearson_bp_r": r_pb,
        "pearson_bp_p": p_pb,
        "pearson_bvrp_r": r_pv,
        "pearson_bvrp_p": p_pv,
        "spearman_bp_r": r_sb,
        "spearman_bp_p": p_sb,
        "spearman_bvrp_r": r_sv,
        "spearman_bvrp_p": p_sv,
        "pearson_bp_cell": _fmt_ic_cell(r_pb, p_pb),
        "pearson_bvrp_cell": _fmt_ic_cell(r_pv, p_pv),
        "spearman_bp_cell": _fmt_ic_cell(r_sb, p_sb),
        "spearman_bvrp_cell": _fmt_ic_cell(r_sv, p_sv),
    }


def build_ic_rows(
    bp: pd.DataFrame,
    bvrp: pd.DataFrame,
    observable_files: list[Path],
    *,
    negate_bvrp_for_display: bool,
) -> list[dict]:
    used_labels: set[str] = set()
    rows: list[dict] = []
    for obs_path in observable_files:
        try:
            obs, freq = load_observable(obs_path)
        except Exception:
            continue
        vcols = observable_value_columns(obs)
        if not vcols:
            continue
        try:
            merged = merge_for_plot(obs, freq, bp, bvrp)
        except Exception:
            continue
        if merged.empty or len(merged) < 3:
            continue
        for vcol in vcols:
            if row := _ic_row_for_column(
                obs_path,
                merged,
                vcol,
                used_labels,
                negate_bvrp_for_display=negate_bvrp_for_display,
            ):
                rows.append(row)
    rows.sort(key=lambda d: str(d["label"]).lower())
    return rows


def _write_ic_excel_block(
    ws,
    start_row: int,
    panel_title: str,
    cols: list[str],
    pearson_bp: list[str],
    pearson_bv: list[str],
    fonts: dict,
    align: dict,
) -> int:
    """写一个 Panel（标题 + 表头 + BP/BVRP 两行），返回下一可用行号。"""
    rr = start_row
    ws.cell(rr, 1, panel_title).font = fonts["title"]
    ws.cell(rr, 1).alignment = align["left"]
    rr += 1
    ws.cell(rr, 1, "")
    for j, name in enumerate(cols, start=2):
        c = ws.cell(rr, j, name)
        c.font = fonts["header"]
        c.alignment = align["right"]
    rr += 1
    ws.cell(rr, 1, "BP").font = fonts["body"]
    ws.cell(rr, 1).alignment = align["left"]
    for j, val in enumerate(pearson_bp, start=2):
        c = ws.cell(rr, j, val)
        c.font = fonts["body"]
        c.alignment = align["right"]
    rr += 1
    ws.cell(rr, 1, "BVRP").font = fonts["body"]
    ws.cell(rr, 1).alignment = align["left"]
    for j, val in enumerate(pearson_bv, start=2):
        c = ws.cell(rr, j, val)
        c.font = fonts["body"]
        c.alignment = align["right"]
    return rr + 1


def export_ic_xlsx(
    bp: pd.DataFrame,
    bvrp: pd.DataFrame,
    run: BoundsRunConfig,
    *,
    out_path: Path | None = None,
) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("[IC 表] 需要 openpyxl（pip install openpyxl）")
        return None

    rows = build_ic_rows(
        bp,
        bvrp,
        list_observable_files(),
        negate_bvrp_for_display=run.negate_bvrp_for_display,
    )
    if not rows:
        print(f"[{run.key}][IC 表] 无有效指标，跳过写入")
        return None

    path = out_path or run.ic_xlsx_path
    run.output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "IC"
    fonts = {
        "title": Font(bold=True, name="Times New Roman", size=11),
        "header": Font(italic=True, name="Times New Roman", size=11),
        "body": Font(name="Times New Roman", size=11),
        "note": Font(italic=True, name="Times New Roman", size=10),
    }
    align = {
        "right": Alignment(horizontal="right", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
    }

    col_labels = [r["label"] for r in rows]
    r = 1
    r = _write_ic_excel_block(
        ws,
        r,
        "Panel A: Pearson correlation (Information Coefficient)",
        col_labels,
        [x["pearson_bp_cell"] for x in rows],
        [x["pearson_bvrp_cell"] for x in rows],
        fonts,
        align,
    )
    r += 1
    r = _write_ic_excel_block(
        ws,
        r,
        "Panel B: Spearman rank correlation",
        col_labels,
        [x["spearman_bp_cell"] for x in rows],
        [x["spearman_bvrp_cell"] for x in rows],
        fonts,
        align,
    )

    note = (
        "Note: *, **, and *** denote two-sided significance at the 10%, 5%, and 1% levels, "
        f"respectively. Bounds: {run.ic_bounds_note}. {run.ic_bvrp_footnote} "
        "Correlations use merged samples without interpolating BP/BVRP."
    )
    ws.cell(r, 1, note).font = fonts["note"]
    ws.cell(r, 1).alignment = align["left"]

    wraw = wb.create_sheet("IC_numeric")
    for j, h in enumerate(IC_NUMERIC_HEADERS, start=1):
        wraw.cell(1, j, h).font = Font(bold=True, name="Times New Roman", size=10)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(IC_NUMERIC_HEADERS, start=1):
            wraw.cell(i, j, row[h])

    wb.save(path)
    return path


def maybe_export_ic_xlsx(
    bp: pd.DataFrame, bvrp: pd.DataFrame, run: BoundsRunConfig
) -> None:
    if not EXPORT_IC_XLSX:
        return
    try:
        if (icp := export_ic_xlsx(bp, bvrp, run)) is not None:
            print(f"[{run.key}] 已保存 IC 表: {icp}")
    except Exception as e:
        print(f"[{run.key}][IC 表] 失败: {e}")


# =============================================================================
# 入口
# =============================================================================
def main() -> None:
    os.chdir(BASE_DIR)
    ensure_derived_observables()

    files = list_observable_files()
    if not files:
        raise FileNotFoundError(f"未找到 CSV/XLSX: {OBSERVABLES_DIR}")

    for run in BOUNDS_RUNS:
        if not bounds_run_inputs_ready(run):
            if run.combined_bp_bvrp_path is not None:
                print(f"[跳过] {run.key}: 未找到 {run.combined_bp_bvrp_path}")
            else:
                print(
                    f"[跳过] {run.key}: 缺少 BP/BVRP CSV，请检查:\n"
                    f"  {run.bp_csv}\n  {run.bvrp_csv}"
                )
            continue
        run.output_dir.mkdir(parents=True, exist_ok=True)
        bp, bvrp = load_bounds_for_run(run)
        plot_all_observables_vs_bounds(files, bp, bvrp, run)
        maybe_export_ic_xlsx(bp, bvrp, run)


if __name__ == "__main__":
    main()
